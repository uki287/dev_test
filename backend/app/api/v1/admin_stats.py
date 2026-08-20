# ============================================================
# 文件功能：数据统计接口（Phase D 扩展）
# 接口（需 stat:view 权限）：
#   GET /admin/stats/dashboard      工作台概览（登录即可见）
#   GET /admin/stats/overview       统计概览卡（各表计数 + 今日动态）
#   GET /admin/stats/pv-trend       访问 PV 趋势（?days=7|30）
#   GET /admin/stats/top-pages      PV Top10 页面
#   GET /admin/stats/aggregate      预约/留言按 类型/状态/时间 聚合
#   GET /admin/stats/export         报表导出 xlsx（PV + 状态分布）
# 权威依据：实施方案 Phase D（数据统计：概览卡、PV 趋势 + Top10、聚合、报表导出）。
# ============================================================
from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.deps import CurrentUser, get_current_user, get_db, require_perm
from app.models.business import Appointment, Message, VisitStat
from app.models.content import Job, News, Product
from app.schemas.response import ApiResp

router = APIRouter(prefix="/admin/stats", tags=["后台-统计"])


@router.get("/dashboard", response_model=ApiResp)
def dashboard(
    _: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """工作台概览：4 统计卡计数 + 近 7 天 PV 序列。"""
    # 1) 4 项业务计数（均排除软删）
    counts = {
        "appointment": db.query(Appointment).filter(Appointment.deleted_at.is_(None)).count(),
        "message": db.query(Message).filter(Message.deleted_at.is_(None)).count(),
        "product": db.query(Product).filter(Product.deleted_at.is_(None)).count(),
        "news": db.query(News).filter(News.deleted_at.is_(None)).count(),
    }
    # 2) 近 7 天 PV：按 stat_date 聚合（缺失日期补 0）
    today = date.today()
    start = today - timedelta(days=6)
    rows = (
        db.query(VisitStat.stat_date, func.sum(VisitStat.pv).label("pv"))
        .filter(VisitStat.stat_date >= start)
        .group_by(VisitStat.stat_date)
        .all()
    )
    pv_map = {str(r.stat_date): int(r.pv or 0) for r in rows}
    pv7 = []
    for i in range(7):
        d = start + timedelta(days=i)
        pv7.append({"date": d.isoformat(), "pv": pv_map.get(d.isoformat(), 0)})
    return ApiResp.ok(data={**counts, "pv7": pv7})


# ---------------- Phase D 扩展（stat:view） ----------------


def _pv_series(db: Session, days: int) -> list[dict]:
    """按天聚合 PV 序列（缺失日期补 0）。"""
    today = date.today()
    start = today - timedelta(days=days - 1)
    rows = (
        db.query(VisitStat.stat_date, func.sum(VisitStat.pv).label("pv"))
        .filter(VisitStat.stat_date >= start)
        .group_by(VisitStat.stat_date)
        .all()
    )
    pv_map = {str(r.stat_date): int(r.pv or 0) for r in rows}
    return [
        {"date": (start + timedelta(days=i)).isoformat(), "pv": pv_map.get((start + timedelta(days=i)).isoformat(), 0)}
        for i in range(days)
    ]


@router.get("/overview", response_model=ApiResp)
def overview(
    _: CurrentUser = Depends(require_perm("stat:view")),
    db: Session = Depends(get_db),
):
    """统计概览卡：各表总数 + 今日新增预约/留言 + 今日 PV。"""
    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    return ApiResp.ok(data={
        "appointment": db.query(Appointment).filter(Appointment.deleted_at.is_(None)).count(),
        "message": db.query(Message).filter(Message.deleted_at.is_(None)).count(),
        "product": db.query(Product).filter(Product.deleted_at.is_(None)).count(),
        "news": db.query(News).filter(News.deleted_at.is_(None)).count(),
        "job": db.query(Job).filter(Job.deleted_at.is_(None)).count(),
        "today_appointment": db.query(Appointment).filter(
            Appointment.deleted_at.is_(None), Appointment.created_date >= today_start).count(),
        "today_message": db.query(Message).filter(
            Message.deleted_at.is_(None), Message.created_date >= today_start).count(),
        "today_pv": db.query(func.coalesce(func.sum(VisitStat.pv), 0))
            .filter(VisitStat.stat_date == today).scalar() or 0,
    })


@router.get("/pv-trend", response_model=ApiResp)
def pv_trend(
    days: int = Query(30, ge=7, le=90),
    _: CurrentUser = Depends(require_perm("stat:view")),
    db: Session = Depends(get_db),
):
    """访问 PV 趋势（7/30/90 天）。"""
    return ApiResp.ok(data=_pv_series(db, days))


@router.get("/top-pages", response_model=ApiResp)
def top_pages(
    days: int = Query(7, ge=1, le=90),
    _: CurrentUser = Depends(require_perm("stat:view")),
    db: Session = Depends(get_db),
):
    """PV Top10 页面（近 N 天）。"""
    start = date.today() - timedelta(days=days - 1)
    rows = (
        db.query(VisitStat.path, func.sum(VisitStat.pv).label("pv"))
        .filter(VisitStat.stat_date >= start)
        .group_by(VisitStat.path)
        .order_by(func.sum(VisitStat.pv).desc())
        .limit(10)
        .all()
    )
    return ApiResp.ok(data=[{"path": r[0], "pv": int(r[1] or 0)} for r in rows])


@router.get("/aggregate", response_model=ApiResp)
def aggregate(
    kind: str = Query("appointment", pattern="^(appointment|message)$", description="业务对象"),
    dimension: str = Query("status", pattern="^(status|type)$", description="聚合维度"),
    _: CurrentUser = Depends(require_perm("stat:view")),
    db: Session = Depends(get_db),
):
    """预约/留言按状态或类型聚合计数。"""
    model = Appointment if kind == "appointment" else Message
    col = model.status if dimension == "status" else model.type
    rows = (
        db.query(col, func.count(model.id))
        .filter(model.deleted_at.is_(None))
        .group_by(col)
        .all()
    )
    return ApiResp.ok(data=[{"key": r[0], "count": r[1]} for r in rows])


@router.get("/export", response_model=None)
def export_report(
    days: int = Query(30, ge=7, le=90),
    _: CurrentUser = Depends(require_perm("stat:view")),
    db: Session = Depends(get_db),
):
    """报表导出 xlsx：近 N 天 PV 趋势 + 预约/留言状态分布。"""
    wb = Workbook()
    # Sheet1：PV 趋势
    ws1 = wb.active
    ws1.title = "PV趋势"
    ws1.append(["日期", "PV"])
    for row in _pv_series(db, days):
        ws1.append([row["date"], row["pv"]])
    # Sheet2：预约状态分布
    ws2 = wb.create_sheet("预约状态")
    ws2.append(["状态", "数量"])
    appt_rows = (
        db.query(Appointment.status, func.count(Appointment.id))
        .filter(Appointment.deleted_at.is_(None))
        .group_by(Appointment.status)
        .all()
    )
    for r in appt_rows:
        ws2.append([r[0], r[1]])
    # Sheet3：留言状态分布
    ws3 = wb.create_sheet("留言状态")
    ws3.append(["状态", "数量"])
    msg_rows = (
        db.query(Message.status, func.count(Message.id))
        .filter(Message.deleted_at.is_(None))
        .group_by(Message.status)
        .all()
    )
    for r in msg_rows:
        ws3.append([r[0], r[1]])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''stats-report.xlsx"},
    )
